from flask import current_app, send_file
from openpyxl import load_workbook
import os
from website import db
import decimal
from sqlalchemy import func
from website.models.settings import Settings
from website.models.purchase import PurchaseItem, PurchaseTax
from website.models.product import Product
from website.jsonify.purchase import getGroupedByAliasTaxes
from datetime import datetime, date, time
from sqlalchemy.exc import IntegrityError


Decimal = decimal.Decimal
FOURPLACES = Decimal(10) ** -4
TWOPLACES = Decimal(10) ** -2


def add_product_to_purchase(product, purchase, current_time):
    product_stock = Decimal(product["stock"]).quantize(FOURPLACES)
    (
        tax_amount,
        product_purchased_price,
        product_purchased_price_wo_tax,
    ) = calculate_tax_and_price(product)
    (product_total_amount, total_tax_amount, total_wo_tax_value) = calculate_totals(
        tax_amount, product_purchased_price, product
    )

    expiration_date = get_expiration_date(product)

    product_query = Product.query.filter_by(name=product["product_name"]).first()
    measure = str(product["measure"])

    found_with_barcode = Product.query.filter_by(barcode=product["barcode"]).first()
    if found_with_barcode and found_with_barcode.name != product["product_name"]:
        if product_query and product_query.barcode != found_with_barcode.barcode:
            raise IntegrityError("barcodeExistsDetailed")

    if product_query:
        purchase_item = create_purchase_item(
            purchase,
            product_query,
            product,
            current_time,
            product_purchased_price,
            total_tax_amount,
            product_total_amount,
            measure,
            product_stock,
            expiration_date,
        )
        product_query.name = product["product_name"]
        product_query.barcode = product["barcode"]
        product_query.tax = product["tax"]
        product_query.purchased_price_wo_tax = product_purchased_price_wo_tax
        product_query.purchased_price = product_purchased_price
        product_query.selling_price = product["selling_price"]
        product_query.measure = measure
        product_query.stock += product_stock
        product_query.expiration_date = expiration_date
    elif found_with_barcode:
        purchase_item = create_purchase_item(
            purchase,
            found_with_barcode,
            product,
            current_time,
            product_purchased_price,
            total_tax_amount,
            product_total_amount,
            measure,
            product_stock,
            expiration_date,
        )
        found_with_barcode.name = product["product_name"]
        found_with_barcode.barcode = product["barcode"]
        found_with_barcode.tax = product["tax"]
        found_with_barcode.purchased_price_wo_tax = product_purchased_price_wo_tax
        found_with_barcode.purchased_price = product_purchased_price
        found_with_barcode.selling_price = product["selling_price"]
        found_with_barcode.measure = measure
        found_with_barcode.stock += product_stock
        found_with_barcode.expiration_date = expiration_date
    else:
        created_product = create_product(
            product, current_time, measure, expiration_date, product_stock
        )

        purchase_item = create_purchase_item(
            purchase,
            created_product,
            product,
            current_time,
            product_purchased_price,
            total_tax_amount,
            product_total_amount,
            measure,
            product_stock,
            expiration_date,
        )

    tax_query = Settings.query.filter_by(settings_value=int(product["tax"])).one()

    tax_value = Decimal(tax_amount * product_stock).quantize(FOURPLACES)
    total_wo_tax_value = product_total_amount - tax_value

    db.session.add(
        PurchaseTax(
            purchase=purchase,
            tax_name=tax_query.settings_name,
            tax_alias=tax_query.settings_alias,
            tax_value=tax_value,
            total_without_tax=total_wo_tax_value,
            date_created=purchase.date_created,
            date_modified=current_time,
        )
    )
    db.session.add(purchase_item)


def get_expiration_date(product):
    if product["expiration_date"]:
        expiration_date = product["expiration_date"].split("-")
        expiration_date = datetime.combine(
            date(
                year=int(expiration_date[0]),
                month=int(expiration_date[1]),
                day=int(expiration_date[2]),
            ),
            time.min,
        )
    else:
        expiration_date = None
    return expiration_date


def create_purchase_item(
    purchase,
    product_query,
    product,
    current_time,
    product_purchased_price,
    tax_amount,
    product_total_amount,
    measure,
    product_stock,
    expiration_date,
):
    return PurchaseItem(
        purchase=purchase,
        product_id=product_query.id,
        product_barcode=product["barcode"],
        product_name=product["product_name"],
        product_tax=product["tax"],
        rabat=product["rabat"],
        product_purchased_price_wo_tax=product["purchased_price_wo_tax"],
        product_purchased_price=product_purchased_price,
        product_selling_price=product["selling_price"],
        product_stock=product_stock,
        product_measure=measure,
        product_expiration_date=expiration_date,
        tax_amount=tax_amount,
        total_amount=product_total_amount,
        date_created=purchase.date_created,
        date_modified=current_time,
    )


def create_product(product, current_time, measure, expiration_date, product_stock):
    product_purchased_price = Decimal(product["purchased_price_wo_tax"])
    created_product = Product(
        name=product["product_name"],
        barcode=product["barcode"],
        stock=product_stock,
        tax=product["tax"],
        purchased_price_wo_tax=product["purchased_price_wo_tax"],
        purchased_price=product_purchased_price,
        selling_price=product["selling_price"],
        measure=measure,
        expiration_date=expiration_date,
        date_created=current_time,
        date_modified=current_time,
    )

    db.session.add(created_product)
    return created_product


def calculate_totals(tax_amount, product_purchased_price, product):
    product_stock = Decimal(product["stock"]).quantize(FOURPLACES)
    product_total_amount = Decimal(product_purchased_price * product_stock).quantize(
        FOURPLACES
    )

    total_tax_amount = Decimal(tax_amount * product_stock).quantize(FOURPLACES)
    total_wo_tax_value = product_total_amount - total_tax_amount

    return product_total_amount, total_tax_amount, total_wo_tax_value


def calculate_tax_and_price(product):
    product_purchased_price_wo_tax = Decimal(product["purchased_price_wo_tax"])
    rabat_percentage = Decimal(product["rabat"]) / 100
    price_wo_rabat = product_purchased_price_wo_tax - (
        product_purchased_price_wo_tax * rabat_percentage
    )

    tax_percentage = Decimal(product["tax"]) / 100
    tax_amount = Decimal(price_wo_rabat * tax_percentage).quantize(FOURPLACES)
    product_purchased_price = Decimal(price_wo_rabat + tax_amount).quantize(FOURPLACES)

    return tax_amount, product_purchased_price, product_purchased_price_wo_tax


def purchase_excel_template(items, file_name):
    try:
        template_path = os.path.join(
            current_app.static_folder, "purchase-template.xlsx"
        )

        wb = load_workbook(template_path)
        ws = wb.active

        rows = 0

        for index, item in enumerate(items):
            rows += 1
            ws[f"A{index+5}"] = index + 1
            ws[f"B{index+5}"] = item.date_created.strftime("%d.%m.%Y")
            ws[f"C{index+5}"] = item.seller_invoice_number
            ws[f"D{index+5}"] = item.seller_name
            ws[f"E{index+5}"] = item.seller_fiscal_number
            ws[f"F{index+5}"] = item.seller_tax_number
            taxes = getGroupedByAliasTaxes(item.purchase_taxes)

            ws[f"G{index+5}"] = (
                taxes["zero"]["total_without_tax"] if "zero" in taxes else 0
            )

            ws[f"M{index+5}"] = (
                taxes["eighteen"]["total_without_tax"] if "eighteen" in taxes else 0
            )
            ws[f"S{index+5}"] = (
                taxes["eighteen"]["tax_value"] if "eighteen" in taxes else 0
            )

            ws[f"V{index+5}"] = (
                taxes["eight"]["total_without_tax"] if "eight" in taxes else 0
            )
            ws[f"AB{index+5}"] = taxes["eight"]["tax_value"] if "eight" in taxes else 0

            ws[f"AC{index+5}"] = f"=SUM(S{index+5},AB{index+5})"

        ws["G4"] = f"=SUM(G5:G{rows + 5})"
        ws["M4"] = f"=SUM(M5:M{rows + 5})"
        ws["S4"] = f"=SUM(S5:S{rows + 5})"
        ws["V4"] = f"=SUM(V5:V{rows + 5})"
        ws["AB4"] = f"=SUM(AB5:AB{rows + 5})"
        ws["AC4"] = f"=SUM(AC5:AC{rows + 5})"

        edited_file_path = os.path.join(current_app.static_folder, "edited_file.xlsx")
        wb.save(edited_file_path)

        return send_file(
            edited_file_path,
            as_attachment=True,
            attachment_filename=f"{file_name}.xlsx",
        )
    except Exception as e:
        return str(e), 500

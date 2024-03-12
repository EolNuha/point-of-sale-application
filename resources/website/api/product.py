from flask import jsonify
from website.models.product import Product
from website.jsonify.product import getProductsList, getProductsExcelList
from website.helpers import get_page_range
from datetime import datetime, date, time
from website import db
from sqlalchemy import or_, asc, desc
from website.api_models.product import (
    product_model,
    product_create_model,
    products_delete_model,
    product_excel_model,
)
from website.api_models.pagination import pagination
from flask_restx import Namespace, Resource, reqparse
from werkzeug.exceptions import BadRequest
from openpyxl import load_workbook
from werkzeug.datastructures import FileStorage
import decimal

product_rest = Namespace("Product")
parser = reqparse.RequestParser()
file_parser = reqparse.RequestParser()
parser.add_argument("page", type=int, default=1)
parser.add_argument("per_page", type=int, default=20)
parser.add_argument("search", type=str, default="")
parser.add_argument("sort_column", type=str, default="id")
parser.add_argument("sort_dir", type=str, default="desc")
file_parser.add_argument(
    "file",
    type=FileStorage,
    location="files",
    required=False,
    help="File upload",
)

Decimal = decimal.Decimal
FOURPLACES = Decimal(10) ** -4
TWOPLACES = Decimal(10) ** -2


@product_rest.route("products")
class ProductsClass(Resource):
    @product_rest.doc(
        params={
            "page": "",
            "per_page": "",
            "search": "",
            "sort_column": "",
            "sort_dir": "",
        }
    )
    @product_rest.marshal_with(pagination)
    def get(self):
        args = parser.parse_args()
        page = args["page"]
        per_page = args["per_page"]
        search = args["search"]
        sort_column = args["sort_column"]
        sort_dir = args["sort_dir"]

        sort = asc(sort_column) if sort_dir == "asc" else desc(sort_column)

        looking_for = (
            search.strip().replace("_", "__").replace("*", "%").replace("?", "_")
            if "*" in search or "_" in search
            else f"%{search.strip()}%"
        )

        paginated_items = (
            Product.query.filter(
                or_(
                    Product.name.ilike(looking_for),
                    Product.id.ilike(looking_for),
                    Product.barcode.ilike(looking_for),
                )
            )
            .order_by(sort)
            .paginate(page=page, per_page=per_page)
        )

        response = {
            "has_next": paginated_items.has_next,
            "has_prev": paginated_items.has_prev,
            "page": paginated_items.page,
            "per_page": paginated_items.per_page,
            "pages": paginated_items.pages,
            "next_num": paginated_items.next_num,
            "prev_num": paginated_items.prev_num,
            "data": getProductsList(paginated_items.items),
            "items": len(paginated_items.items),
            "total": paginated_items.total,
            "page_range": get_page_range(paginated_items.page, paginated_items.pages),
        }
        return response, 200

    @product_rest.expect(product_create_model)
    @product_rest.marshal_with(product_model)
    @product_rest.doc(responses={200: "Success", 500: "Server Error"})
    def post(self):
        name = product_rest.payload["name"]
        barcode = product_rest.payload["barcode"]
        stock = product_rest.payload["stock"]
        tax = product_rest.payload["tax"]
        purchased_price_wo_tax = product_rest.payload["purchased_price_wo_tax"]
        purchased_price = product_rest.payload["purchased_price"]
        selling_price = product_rest.payload["selling_price"]
        expiration_date = product_rest.payload["expiration_date"]
        measure = product_rest.payload["measure"]

        if expiration_date:
            expiration_date = expiration_date.split("-")
            expiration_date = datetime.combine(
                date(
                    year=int(expiration_date[0]),
                    month=int(expiration_date[1]),
                    day=int(expiration_date[2]),
                ),
                time.min,
            )
        else:
            expiration_date = None

        found_with_name = Product.query.filter_by(name=name).first()
        found_with_barcode = Product.query.filter_by(barcode=barcode).first()

        if found_with_name:
            raise BadRequest("productWithNameExists")
        if found_with_barcode:
            raise BadRequest("productWithBarcodeExists")

        product = Product(
            name=name,
            barcode=barcode,
            stock=stock,
            tax=tax,
            purchased_price_wo_tax=purchased_price_wo_tax,
            purchased_price=purchased_price,
            selling_price=selling_price,
            measure=measure,
            expiration_date=expiration_date,
            date_created=datetime.now(),
            date_modified=datetime.now(),
        )

        db.session.add(product)
        db.session.commit()
        return product, 200

    @product_rest.expect(products_delete_model)
    def delete(self):
        products = product_rest.payload["products"]
        Product.query.filter(Product.id.in_(products)).delete()
        db.session.commit()
        return "Success", 200


@product_rest.route("products/excel")
class GetProductsExcel(Resource):
    @product_rest.doc(
        params={
            "search": "",
            "sort_column": "",
            "sort_dir": "",
        }
    )
    @product_rest.marshal_with(product_excel_model)
    def get(self):
        args = parser.parse_args()
        search = args["search"]
        sort_column = args["sort_column"]
        sort_dir = args["sort_dir"]

        sort = asc(sort_column) if sort_dir == "asc" else desc(sort_column)

        looking_for = (
            search.strip().replace("_", "__").replace("*", "%").replace("?", "_")
            if "*" in search or "_" in search
            else f"%{search.strip()}%"
        )

        items = (
            Product.query.filter(
                or_(
                    Product.name.ilike(looking_for),
                    Product.id.ilike(looking_for),
                    Product.barcode.ilike(looking_for),
                )
            )
            .order_by(sort)
            .all()
        )

        return getProductsExcelList(items), 200

    @product_rest.expect(file_parser)
    def put(self):
        args = file_parser.parse_args()
        file = args["file"]

        wb = load_workbook(file)
        ws = wb.active

        arr = [
            "Barkodi",
            "Emri",
            "Sasia",
            "TVSH",
            "Çmimi i blerjes",
            "Çmimi i shitjes",
            "Njesia matese",
        ]

        if ws.max_row > 501:
            raise BadRequest("Numri maksimal i produkteve eshte 500!")

        for index, row in enumerate(ws.iter_rows(values_only=True)):
            if index > 0:
                break
            for idx, value in enumerate(row):
                if value.lower() != arr[idx].lower():
                    raise BadRequest(f'Kolona "{arr[idx]}" eshte gabim!')

        if ws.max_row == 1:
            raise BadRequest("Asnje produkt nuk eshte shtuar!")

        for index, row in enumerate(ws.iter_rows(values_only=True)):
            if index == 0:
                continue
            self.validate_inputs(row)
            try:
                purchased_price_wo_tax = Decimal(row[4])
                tax_percentage = Decimal(row[3]) / 100
                tax_amount = Decimal(purchased_price_wo_tax * tax_percentage).quantize(
                    FOURPLACES
                )
                product_purchased_price = Decimal(
                    purchased_price_wo_tax + tax_amount
                ).quantize(FOURPLACES)

                product = Product(
                    barcode=row[0],
                    name=row[1],
                    stock=row[2],
                    tax=float(row[3]),
                    purchased_price_wo_tax=purchased_price_wo_tax,
                    purchased_price=product_purchased_price,
                    selling_price=Decimal(row[5]).quantize(FOURPLACES),
                    measure=row[6],
                    date_created=datetime.now(),
                    date_modified=datetime.now(),
                )

                db.session.add(product)
            except Exception as e:
                print(e)
                return "Something went wrong!", 500

        db.session.commit()

        return jsonify(arr)

    def validate_inputs(self, row):
        arr = [
            "Barkodi",
            "Emri",
            "Sasia",
            "TVSH",
            "Çmimi i blerjes",
            "Çmimi i shitjes",
            "Njesia matese",
        ]
        for i in range(0, 7):
            if not row[i] and row[i] != 0:
                raise BadRequest(
                    f"Fusha {arr[i]} {row[i]} {type(row[i])} nuk mund te jete e zbrazet!"
                )

            if i not in [1, 6]:
                try:
                    if i == 3 and float(row[i]) not in [0, 8, 18]:
                        raise BadRequest("Jo Valid!")
                    else:
                        float(row[i])
                except Exception as e:
                    raise BadRequest(f"Vlera {row[i]} nuk eshte {arr[i]} valid!")

            if i == 6 and row[i].lower() not in ["kg", "pcs", "liter"]:
                raise BadRequest(f"Njesia matese duhet te jete pcs, kg, ose liter")


@product_rest.route("products/<int:productId>")
class ProductDetails(Resource):
    @product_rest.marshal_with(product_model)
    def get(self, productId):
        return Product.query.filter_by(id=productId).first_or_404()

    @product_rest.expect(product_create_model)
    @product_rest.marshal_with(product_model)
    def put(self, productId):
        name = product_rest.payload["name"]
        barcode = product_rest.payload["barcode"]
        stock = product_rest.payload["stock"]
        tax = product_rest.payload["tax"]
        purchased_price_wo_tax = product_rest.payload["purchased_price_wo_tax"]
        purchased_price = product_rest.payload["purchased_price"]
        selling_price = product_rest.payload["selling_price"]
        expiration_date = product_rest.payload["expiration_date"]
        measure = product_rest.payload["measure"]

        if expiration_date:
            expiration_date = expiration_date[0:10].split("-")
            expiration_date = datetime.combine(
                date(
                    year=int(expiration_date[0]),
                    month=int(expiration_date[1]),
                    day=int(expiration_date[2]),
                ),
                time.min,
            )
        else:
            expiration_date = None

        product = Product.query.filter_by(id=productId).first_or_404()

        found_with_name = Product.query.filter_by(name=name).first()
        found_with_barcode = Product.query.filter_by(barcode=barcode).first()

        if found_with_name and name != product.name:
            raise BadRequest("productWithNameExists")
        if found_with_barcode and barcode != product.barcode:
            raise BadRequest("productWithBarcodeExists")

        product.name = name
        product.barcode = barcode
        product.stock = stock
        product.tax = tax
        product.purchased_price_wo_tax = purchased_price_wo_tax
        product.purchased_price = purchased_price
        product.selling_price = selling_price
        product.measure = measure
        product.expiration_date = expiration_date
        product.date_modified = datetime.now()

        db.session.commit()
        return product, 200

    def delete(self, productId):
        Product.query.filter_by(id=productId).delete()
        db.session.commit()
        return "Success", 200


@product_rest.route("products/barcode/<int:barcode>")
class getProductDetailsByBarcode(Resource):
    @product_rest.marshal_with(product_model)
    def get(self, barcode):
        return Product.query.filter_by(barcode=barcode).first_or_404(), 200

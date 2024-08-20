from flask import current_app, send_file
from openpyxl import load_workbook
from website.jsonify.purchase import getGroupedByAliasTaxes
import os

def sale_excel_template(items, file_name):
    try:
        template_path = os.path.join(
            current_app.static_folder, "sale-template.xlsx"
        )

        wb = load_workbook(template_path)
        ws = wb.active

        rows = 0

        for index, item in enumerate(items):
            rows += 1
            ws[f"A{index+4}"] = index + 1
            ws[f"B{index+4}"] = item.date_created.strftime("%d.%m.%Y")
            ws[f"C{index+4}"] = item.id
            ws[f"D{index+4}"] = "Qytetar"
            # ws[f"E{index+4}"] = "e"
            # ws[f"F{index+4}"] = "e"
            taxes = getGroupedByAliasTaxes(item.sale_taxes)
            print(taxes, item.id, item.date_created.strftime("%d.%m.%Y"), index+4)

            ws[f"G{index+4}"] = (
                taxes["zero"]["total_without_tax"] if "zero" in taxes else 0
            )

            ws[f"M{index+4}"] = (
                taxes["eighteen"]["total_without_tax"] if "eighteen" in taxes else 0
            )
            ws[f"R{index+4}"] = (
                taxes["eighteen"]["tax_value"] if "eighteen" in taxes else 0
            )

            ws[f"S{index+4}"] = (
                taxes["eight"]["total_without_tax"] if "eight" in taxes else 0
            )
            ws[f"W{index+4}"] = taxes["eight"]["tax_value"] if "eight" in taxes else 0

            ws[f"X{index+4}"] = f"=SUM(R{index+4},W{index+4})"

        # ws["G4"] = f"=SUM(G4:G{rows + 4})"
        # ws["M4"] = f"=SUM(M4:M{rows + 4})"
        # ws["S4"] = f"=SUM(S4:S{rows + 4})"
        # ws["V4"] = f"=SUM(V4:V{rows + 4})"
        # ws["AB4"] = f"=SUM(AB4:AB{rows + 4})"
        # ws["AC4"] = f"=SUM(AC4:AC{rows + 4})"

        edited_file_path = os.path.join(current_app.static_folder, "edited_file.xlsx")
        wb.save(edited_file_path)

        return send_file(
            edited_file_path,
            as_attachment=True,
            attachment_filename=f"{file_name}.xlsx",
        )
    except Exception as e:
        return str(e), 500